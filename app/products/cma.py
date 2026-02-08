"""Product 1 — CMA Engine: Comparative Market Analysis in ~5 minutes."""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any

from rich.console import Console
from rich.panel import Panel
from rich.prompt import Confirm, Prompt
from rich.table import Table

from app.config import PROMPTS_DIR, TEMPLATES_DIR
from app.models import claude, gemini
from app.utils import deal as deal_mgr
from app.utils.spreadsheet import (
    CURRENCY_FORMAT,
    PERCENT_FORMAT,
    create_workbook,
    save_workbook,
    write_kv_pairs,
    write_section_header,
    write_table,
)

console = Console()


# ── helpers ──────────────────────────────────────────────────────────────


def _collect_manual_comps() -> list[dict[str, Any]]:
    """Interactively collect comparable sales from the user."""
    comps: list[dict[str, Any]] = []
    console.print("\n[bold]Enter comparable sales[/bold] (type 'done' when finished):\n")

    while True:
        console.print(f"[cyan]─── Comp #{len(comps) + 1} ───[/cyan]")
        address = Prompt.ask("  Address (or 'done')")
        if address.lower() == "done":
            break

        sold_price = Prompt.ask("  Sold price ($)", default="0")
        sold_date = Prompt.ask("  Sold date (YYYY-MM-DD)", default="")
        beds = Prompt.ask("  Bedrooms", default="0")
        baths = Prompt.ask("  Bathrooms", default="0")
        cars = Prompt.ask("  Car spaces", default="0")
        land_sqm = Prompt.ask("  Land area (sqm)", default="0")
        building_sqm = Prompt.ask("  Building area (sqm)", default="0")
        prop_type = Prompt.ask("  Property type", default="house")
        condition = Prompt.ask("  Condition notes", default="")
        distance = Prompt.ask("  Distance from subject (km)", default="0")

        comp = {
            "address": address,
            "sold_price": int(sold_price.replace(",", "").replace("$", "")),
            "sold_date": sold_date,
            "beds": int(beds),
            "baths": int(baths),
            "cars": int(cars),
            "land_sqm": float(land_sqm),
            "building_sqm": float(building_sqm),
            "property_type": prop_type,
            "condition_notes": condition,
            "distance_km": float(distance),
        }
        comps.append(comp)
        console.print(f"  [green]✓ Comp added: {address} — ${comp['sold_price']:,}[/green]\n")

    return comps


def _load_comps_from_file(file_path: str) -> list[dict[str, Any]]:
    """Load comps from a JSON file (exported or manually created)."""
    path = Path(file_path)
    if not path.exists():
        console.print(f"[red]File not found: {file_path}[/red]")
        return []
    data = json.loads(path.read_text())
    if isinstance(data, list):
        return data
    if isinstance(data, dict) and "comps" in data:
        return data["comps"]
    return []


def _display_cma_results(result: dict[str, Any]) -> None:
    """Pretty-print CMA results to console."""
    if "parse_error" in result:
        console.print("[red]⚠ Could not parse Claude response as JSON.[/red]")
        console.print(result.get("raw_response", ""))
        return

    valuation = result.get("valuation", {})
    console.print()
    console.print(Panel(
        f"[bold green]Value Range:[/bold green]  "
        f"${valuation.get('value_range_low', 0):,.0f}  –  "
        f"${valuation.get('value_range_high', 0):,.0f}\n"
        f"[bold]Point Estimate:[/bold]  ${valuation.get('point_estimate', 0):,.0f}\n"
        f"[bold]Confidence:[/bold]  {valuation.get('confidence_score', 0):.0%}\n"
        f"[dim]{valuation.get('confidence_reasoning', '')}[/dim]",
        title="📊 CMA Valuation Result",
        border_style="green",
    ))

    # Comps table
    comps = result.get("comps_analysis", [])
    if comps:
        table = Table(title="Comparable Sales Analysis")
        table.add_column("Address", style="cyan", max_width=30)
        table.add_column("Sold Price", justify="right")
        table.add_column("Adjusted", justify="right")
        table.add_column("Tag", justify="center")
        table.add_column("Weight", justify="center")
        table.add_column("Reasoning", max_width=40)

        for c in comps:
            table.add_row(
                c.get("address", ""),
                f"${c.get('sold_price', 0):,.0f}",
                f"${c.get('adjusted_price', 0):,.0f}",
                c.get("similarity_tag", ""),
                f"{c.get('weight', 0):.0%}",
                c.get("similarity_reasoning", "")[:60],
            )
        console.print(table)


# ── main CMA workflow ────────────────────────────────────────────────────


def run_cma(deal_dir: Path | None = None) -> dict[str, Any]:
    """Run the full CMA workflow interactively."""
    console.print(Panel("[bold]Product 1 — CMA Engine[/bold]\nComparative Market Analysis", border_style="blue"))

    # ── Step 0: Deal setup ───────────────────────────────────────────
    if deal_dir is None:
        address = Prompt.ask("Property address")
        listing_url = Prompt.ask("Listing URL (optional)", default="")
        notes = Prompt.ask("Notes (optional)", default="")
        deal_dir = deal_mgr.create_deal(address, listing_url, notes)
        console.print(f"[green]✓ Deal created:[/green] {deal_dir.name}")
    else:
        meta = deal_mgr.load_deal(deal_dir)
        address = meta["address"]
        console.print(f"[green]✓ Using existing deal:[/green] {address}")

    meta = deal_mgr.load_deal(deal_dir)

    # ── Step 1: Photos ───────────────────────────────────────────────
    console.print("\n[bold]Step 1: Listing Photos[/bold]")
    photos = deal_mgr.list_photos(deal_dir)

    if not photos:
        console.print("No photos found in deal folder.")
        photo_input = Prompt.ask(
            "Enter photo file paths (comma-separated) or 'skip'",
            default="skip",
        )
        if photo_input.lower() != "skip":
            paths = [p.strip().strip('"') for p in photo_input.split(",")]
            photos = deal_mgr.add_photos(deal_dir, paths)
            console.print(f"[green]✓ {len(photos)} photos added[/green]")

    if photos:
        console.print(f"[green]✓ {len(photos)} photos available[/green]")

    # ── Step 2: Gemini vision extraction ─────────────────────────────
    vision_data: dict[str, Any] = {}
    if photos:
        console.print("\n[bold]Step 2: Gemini Vision Extraction[/bold]")
        run_vision = Confirm.ask("Run Gemini vision extraction on photos?", default=True)

        if run_vision:
            prompt_text = (PROMPTS_DIR / "gemini_vision_extraction.md").read_text()
            console.print("[dim]Sending photos to Gemini...[/dim]")
            try:
                vision_data = gemini.extract_listing_facts(photos, prompt_text, address)
                deal_mgr.save_output(deal_dir, "vision_extraction.json", vision_data)
                deal_mgr.save_log(deal_dir, "gemini_vision", json.dumps(vision_data, indent=2))

                if vision_data.get("parse_error"):
                    console.print("[yellow]⚠ Vision response wasn't valid JSON. Raw saved to outputs.[/yellow]")
                else:
                    console.print("[green]✓ Vision extraction complete[/green]")
                    # Show key facts
                    for key in ("property_type", "condition_overall", "finish_level", "bedrooms", "bathrooms"):
                        if key in vision_data:
                            console.print(f"  {key}: [cyan]{vision_data[key]}[/cyan]")
            except Exception as e:
                console.print(f"[red]✗ Gemini error: {e}[/red]")
                console.print("[dim]Continuing without vision data...[/dim]")
    else:
        console.print("\n[dim]Step 2: Skipping vision extraction (no photos)[/dim]")

    # ── Step 3: Comparable sales ─────────────────────────────────────
    console.print("\n[bold]Step 3: Comparable Sales[/bold]")
    console.print("How would you like to provide comps?")
    console.print("  [1] Enter manually")
    console.print("  [2] Load from JSON file")
    console.print("  [3] Both (manual + file)")
    comp_method = Prompt.ask("Choice", choices=["1", "2", "3"], default="1")

    comps: list[dict[str, Any]] = []

    if comp_method in ("1", "3"):
        comps.extend(_collect_manual_comps())

    if comp_method in ("2", "3"):
        file_path = Prompt.ask("Path to comps JSON file")
        file_comps = _load_comps_from_file(file_path)
        comps.extend(file_comps)
        console.print(f"[green]✓ Loaded {len(file_comps)} comps from file[/green]")

    if not comps:
        console.print("[red]No comps provided. Cannot produce CMA.[/red]")
        return {}

    console.print(f"\n[green]✓ {len(comps)} total comparable sales[/green]")
    deal_mgr.save_output(deal_dir, "comps_input.json", comps)

    # ── Step 4: Claude CMA reasoning ─────────────────────────────────
    console.print("\n[bold]Step 4: Claude CMA Analysis[/bold]")
    console.print("[dim]Sending data to Claude for analysis...[/dim]")

    system_prompt = (PROMPTS_DIR / "claude_cma_reasoning.md").read_text()

    user_message = json.dumps({
        "subject_property": {
            "address": address,
            "listing_url": meta.get("listing_url", ""),
            "notes": meta.get("notes", ""),
            "vision_extraction": vision_data,
        },
        "comparable_sales": comps,
    }, indent=2)

    try:
        cma_result = claude.reason(system_prompt, user_message, expect_json=True)
    except Exception as e:
        console.print(f"[red]✗ Claude error: {e}[/red]")
        return {}

    deal_mgr.save_output(deal_dir, "cma_result.json", cma_result)
    deal_mgr.save_log(deal_dir, "claude_cma", json.dumps(cma_result, indent=2, default=str))

    # ── Step 5: Display results ──────────────────────────────────────
    console.print("\n[bold]Step 5: CMA Results[/bold]")
    _display_cma_results(cma_result)

    # ── Step 6: Generate spreadsheet ─────────────────────────────────
    console.print("\n[bold]Step 6: Generating Spreadsheet[/bold]")
    xlsx_path = _generate_cma_spreadsheet(deal_dir, meta, vision_data, comps, cma_result)
    console.print(f"[green]✓ Spreadsheet saved:[/green] {xlsx_path}")

    # ── Step 7: Generate markdown summary ────────────────────────────
    md_path = _generate_cma_markdown(deal_dir, meta, cma_result)
    console.print(f"[green]✓ Markdown saved:[/green] {md_path}")

    # ── Step 8: Human sign-off ───────────────────────────────────────
    console.print()
    signed_off = Confirm.ask(
        "[bold yellow]⚠ HUMAN SIGN-OFF:[/bold yellow] Review the outputs above. Approve this CMA?",
        default=False,
    )
    meta["cma_status"] = "approved" if signed_off else "draft"
    meta["cma_completed_at"] = datetime.now().isoformat()
    deal_mgr.save_deal_meta(deal_dir, meta)

    status = "[green]APPROVED[/green]" if signed_off else "[yellow]DRAFT[/yellow]"
    console.print(f"\nCMA Status: {status}")
    console.print(f"Deal folder: [cyan]{deal_dir}[/cyan]")

    return cma_result


# ── spreadsheet generation ───────────────────────────────────────────────


def _generate_cma_spreadsheet(
    deal_dir: Path,
    meta: dict[str, Any],
    vision_data: dict[str, Any],
    comps: list[dict[str, Any]],
    cma_result: dict[str, Any],
) -> Path:
    """Generate the CMA spreadsheet with all data and calculations."""
    wb = create_workbook()
    ws = wb.active
    ws.title = "CMA Summary"

    row = 1

    # ── Subject Property ─────────────────────────────────────────────
    row = write_section_header(ws, row, "SUBJECT PROPERTY", 8)
    subject = cma_result.get("subject_summary", {})
    kv = [
        ("Address", meta.get("address", "")),
        ("Property Type", subject.get("property_type", vision_data.get("property_type", ""))),
        ("Bedrooms", subject.get("beds", vision_data.get("bedrooms", ""))),
        ("Bathrooms", subject.get("baths", vision_data.get("bathrooms", ""))),
        ("Car Spaces", subject.get("cars", vision_data.get("car_spaces", ""))),
        ("Land Area (sqm)", subject.get("land_sqm", vision_data.get("land_area_sqm", ""))),
        ("Building Area (sqm)", subject.get("building_sqm", vision_data.get("building_area_sqm", ""))),
        ("Condition", subject.get("condition", vision_data.get("condition_overall", ""))),
        ("Finish Level", subject.get("finish_level", vision_data.get("finish_level", ""))),
        ("Listing URL", meta.get("listing_url", "")),
    ]
    row = write_kv_pairs(ws, kv, start_row=row)
    row += 1

    # ── Comps Table ──────────────────────────────────────────────────
    row = write_section_header(ws, row, "COMPARABLE SALES ANALYSIS", 12)
    comps_analysis = cma_result.get("comps_analysis", [])

    comp_headers = [
        "Address", "Sold Price", "Sold Date", "Beds", "Baths", "Cars",
        "Land (sqm)", "Tag", "Adjusted Price", "Weight", "Reasoning",
    ]
    comp_rows = []
    for c in comps_analysis:
        comp_rows.append([
            c.get("address", ""),
            c.get("sold_price", 0),
            c.get("sold_date", ""),
            c.get("beds", ""),
            c.get("baths", ""),
            c.get("cars", ""),
            c.get("land_sqm", ""),
            c.get("similarity_tag", ""),
            c.get("adjusted_price", 0),
            c.get("weight", 0),
            c.get("similarity_reasoning", ""),
        ])

    row = write_table(
        ws, comp_headers, comp_rows, start_row=row,
        col_widths=[30, 14, 12, 6, 6, 6, 10, 10, 14, 8, 40],
        number_formats={1: CURRENCY_FORMAT, 8: CURRENCY_FORMAT, 9: PERCENT_FORMAT},
    )
    row += 1

    # ── Adjustments Detail ───────────────────────────────────────────
    row = write_section_header(ws, row, "ADJUSTMENT DETAILS", 8)
    for c in comps_analysis:
        adj = c.get("adjustments", [])
        if adj:
            from openpyxl.styles import Font as _Font
            ws.cell(row=row, column=1, value=c.get("address", "")).font = _Font(bold=True)
            row += 1
            adj_headers = ["Factor", "Direction", "Amount (%)", "Reasoning"]
            adj_rows = [[a.get("factor", ""), a.get("direction", ""), a.get("amount_pct", 0), a.get("reasoning", "")] for a in adj]
            row = write_table(ws, adj_headers, adj_rows, start_row=row, col_widths=[20, 10, 12, 50])
            row += 1

    # ── Valuation ────────────────────────────────────────────────────
    row = write_section_header(ws, row, "VALUATION", 8)
    val = cma_result.get("valuation", {})
    val_kv = [
        ("Methodology", val.get("methodology", "")),
        ("Weighted Average", val.get("weighted_average", 0)),
        ("Value Range — Low", val.get("value_range_low", 0)),
        ("Value Range — High", val.get("value_range_high", 0)),
        ("Point Estimate", val.get("point_estimate", 0)),
        ("Confidence Score", val.get("confidence_score", 0)),
        ("Confidence Reasoning", val.get("confidence_reasoning", "")),
    ]
    row = write_kv_pairs(ws, val_kv, start_row=row)
    row += 1

    # ── Assumptions & Caveats ────────────────────────────────────────
    row = write_section_header(ws, row, "ASSUMPTIONS & CAVEATS", 8)
    for a in val.get("assumptions", []):
        ws.cell(row=row, column=1, value="Assumption")
        ws.cell(row=row, column=2, value=a)
        row += 1
    for c in val.get("caveats", []):
        ws.cell(row=row, column=1, value="Caveat")
        ws.cell(row=row, column=2, value=c)
        row += 1

    # ── Save ─────────────────────────────────────────────────────────
    xlsx_path = deal_dir / "outputs" / "cma_report.xlsx"
    return save_workbook(wb, xlsx_path)


def _generate_cma_markdown(
    deal_dir: Path,
    meta: dict[str, Any],
    cma_result: dict[str, Any],
) -> Path:
    """Generate a human-readable Markdown CMA summary."""
    val = cma_result.get("valuation", {})
    subject = cma_result.get("subject_summary", {})
    comps = cma_result.get("comps_analysis", [])

    lines = [
        f"# CMA Report — {meta.get('address', 'Unknown')}",
        f"_Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}_",
        "",
        "## Subject Property",
        f"- **Address**: {meta.get('address', '')}",
        f"- **Type**: {subject.get('property_type', '')}",
        f"- **Beds/Baths/Cars**: {subject.get('beds', '?')}/{subject.get('baths', '?')}/{subject.get('cars', '?')}",
        f"- **Land**: {subject.get('land_sqm', '?')} sqm",
        f"- **Condition**: {subject.get('condition', '')}",
        f"- **Finish**: {subject.get('finish_level', '')}",
        "",
        "## Valuation",
        f"- **Value Range**: ${val.get('value_range_low', 0):,.0f} – ${val.get('value_range_high', 0):,.0f}",
        f"- **Point Estimate**: ${val.get('point_estimate', 0):,.0f}",
        f"- **Confidence**: {val.get('confidence_score', 0):.0%}",
        f"- **Methodology**: {val.get('methodology', '')}",
        "",
        "## Comparable Sales",
        "",
        "| Address | Sold Price | Tag | Adjusted | Weight |",
        "|---------|-----------|-----|----------|--------|",
    ]

    for c in comps:
        lines.append(
            f"| {c.get('address', '')} | ${c.get('sold_price', 0):,.0f} | "
            f"{c.get('similarity_tag', '')} | ${c.get('adjusted_price', 0):,.0f} | "
            f"{c.get('weight', 0):.0%} |"
        )

    lines.extend([
        "",
        "## Market Commentary",
        cma_result.get("market_commentary", ""),
        "",
        "## Recommendations",
        cma_result.get("recommendations", ""),
        "",
        "---",
        f"_Status: {meta.get('cma_status', 'draft')}_",
    ])

    md_path = deal_dir / "outputs" / "cma_summary.md"
    md_path.write_text("\n".join(lines))
    return md_path
