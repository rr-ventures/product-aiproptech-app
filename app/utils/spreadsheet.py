"""Spreadsheet generation helpers using openpyxl."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── style constants ──────────────────────────────────────────────────────
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CURRENCY_FORMAT = '"$"#,##0'
PERCENT_FORMAT = "0.0%"
NUMBER_FORMAT = "#,##0"


def create_workbook() -> Workbook:
    """Create a new workbook with default settings."""
    wb = Workbook()
    return wb


def write_table(
    ws,
    headers: list[str],
    rows: list[list[Any]],
    start_row: int = 1,
    start_col: int = 1,
    col_widths: list[int] | None = None,
    number_formats: dict[int, str] | None = None,
) -> int:
    """Write a formatted table to a worksheet. Returns the next available row."""
    # Headers
    for ci, header in enumerate(headers, start=start_col):
        cell = ws.cell(row=start_row, column=ci, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Data rows
    for ri, row_data in enumerate(rows, start=start_row + 1):
        for ci, val in enumerate(row_data, start=start_col):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True)
            if number_formats and (ci - start_col) in number_formats:
                cell.number_format = number_formats[ci - start_col]

    # Column widths
    if col_widths:
        for ci, w in enumerate(col_widths, start=start_col):
            ws.column_dimensions[get_column_letter(ci)].width = w

    return start_row + 1 + len(rows)


def write_section_header(ws, row: int, text: str, col_span: int = 8) -> int:
    """Write a section header row spanning multiple columns."""
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, size=12, color="2F5496")
    cell.fill = SUBHEADER_FILL
    for ci in range(2, col_span + 1):
        ws.cell(row=row, column=ci).fill = SUBHEADER_FILL
    return row + 1


def write_kv_pairs(
    ws,
    pairs: list[tuple[str, Any]],
    start_row: int = 1,
    key_col: int = 1,
    val_col: int = 2,
    val_format: str | None = None,
) -> int:
    """Write key-value pairs in two columns."""
    for i, (key, val) in enumerate(pairs):
        r = start_row + i
        kc = ws.cell(row=r, column=key_col, value=key)
        kc.font = Font(bold=True)
        kc.border = THIN_BORDER
        vc = ws.cell(row=r, column=val_col, value=val)
        vc.border = THIN_BORDER
        if val_format:
            vc.number_format = val_format
    return start_row + len(pairs)


def save_workbook(wb: Workbook, path: Path) -> Path:
    """Save workbook to path, creating parent dirs if needed."""
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    return path
