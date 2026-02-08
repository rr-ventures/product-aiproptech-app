"""Spreadsheet → JSON converter for templates and comp data.

Handles Excel (.xlsx/.xls) and CSV files. Uses smart column name matching
to map whatever columns are in the spreadsheet to the expected JSON structure.
"""

from __future__ import annotations

import csv
import io
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


# ═══════════════════════════════════════════════════════════════════════
# CORE: Read any spreadsheet into rows of dicts
# ═══════════════════════════════════════════════════════════════════════

def read_spreadsheet(file_path: str | Path = None, file_bytes: bytes = None, filename: str = "") -> list[dict[str, Any]]:
    """
    Read a spreadsheet (xlsx or csv) into a list of dicts.
    Accepts either a file path or raw bytes + filename.
    Returns rows as dicts with cleaned column names as keys.
    """
    if file_path:
        path = Path(file_path)
        filename = path.name
        file_bytes = path.read_bytes()

    ext = Path(filename).suffix.lower()

    if ext in (".xlsx", ".xls"):
        return _read_excel(file_bytes)
    elif ext == ".csv":
        return _read_csv(file_bytes)
    else:
        raise ValueError(f"Unsupported file type: {ext}. Use .xlsx, .xls, or .csv")


def _read_excel(data: bytes) -> list[dict[str, Any]]:
    """Read Excel file into list of dicts."""
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    # Find header row (first row with mostly non-empty cells)
    headers = None
    for row in rows_iter:
        non_empty = [c for c in row if c is not None and str(c).strip()]
        if len(non_empty) >= 2:  # At least 2 columns with data = header row
            headers = [_clean_col_name(str(c)) if c else f"col_{i}" for i, c in enumerate(row)]
            break

    if not headers:
        return []

    result = []
    for row in rows_iter:
        if all(c is None or str(c).strip() == "" for c in row):
            continue  # skip empty rows
        row_dict = {}
        for i, val in enumerate(row):
            if i < len(headers):
                row_dict[headers[i]] = _clean_value(val)
        result.append(row_dict)

    wb.close()
    return result


def _read_csv(data: bytes) -> list[dict[str, Any]]:
    """Read CSV file into list of dicts."""
    # Try to decode
    for encoding in ("utf-8", "utf-8-sig", "latin-1", "cp1252"):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = data.decode("utf-8", errors="replace")

    reader = csv.DictReader(io.StringIO(text))
    result = []
    for row in reader:
        cleaned = {_clean_col_name(k): _clean_value(v) for k, v in row.items() if k}
        result.append(cleaned)
    return result


def _clean_col_name(name: str) -> str:
    """Normalize column names: lowercase, strip, replace spaces/special chars."""
    name = str(name).strip().lower()
    name = re.sub(r"[^\w\s]", "", name)
    name = re.sub(r"\s+", "_", name)
    return name


def _clean_value(val: Any) -> Any:
    """Clean cell values."""
    if val is None:
        return ""
    if isinstance(val, str):
        val = val.strip()
        # Try to convert numeric strings
        if val.replace(",", "").replace("$", "").replace("%", "").replace(".", "").replace("-", "").isdigit():
            cleaned = val.replace(",", "").replace("$", "").replace("%", "")
            try:
                if "." in cleaned:
                    return float(cleaned)
                return int(cleaned)
            except ValueError:
                pass
        return val
    return val


# ═══════════════════════════════════════════════════════════════════════
# COLUMN MAPPING: fuzzy-match spreadsheet columns to expected fields
# ═══════════════════════════════════════════════════════════════════════

# For each target field, list likely column name patterns
_DD_CHECKLIST_MAP = {
    "item_number": ["item_number", "number", "no", "num", "id", "#", "item_no", "step"],
    "category": ["category", "cat", "group", "section", "area", "type"],
    "name": ["name", "item", "description", "desc", "check", "checklist_item", "task", "action", "item_name", "item_description"],
    "source": ["source", "data_source", "where", "reference", "ref", "portal", "website", "how"],
    "risk_if_fail": ["risk", "risk_if_fail", "risk_level", "severity", "priority", "importance", "impact"],
    "notes": ["notes", "note", "comments", "comment", "details", "additional", "remarks"],
}

_STORES_MAP = {
    "category": ["category", "cat", "type", "product_type", "section", "area"],
    "name": ["name", "store", "store_name", "supplier", "company", "vendor", "shop"],
    "notes": ["notes", "note", "comments", "description", "details"],
    "website": ["website", "url", "web", "link", "site"],
}

_COMPS_MAP = {
    "address": ["address", "property", "location", "street", "property_address"],
    "sold_price": ["sold_price", "price", "sale_price", "sold", "amount", "sale_amount", "sold_for"],
    "sold_date": ["sold_date", "date", "sale_date", "settlement", "settlement_date", "date_sold"],
    "beds": ["beds", "bedrooms", "bed", "br", "bedroom"],
    "baths": ["baths", "bathrooms", "bath", "bathroom"],
    "cars": ["cars", "car_spaces", "car", "parking", "garage", "carspace"],
    "land_sqm": ["land_sqm", "land", "land_area", "land_size", "lot_size", "sqm", "area"],
    "building_sqm": ["building_sqm", "building", "building_area", "floor_area", "internal", "house_size"],
    "property_type": ["property_type", "type", "prop_type", "dwelling"],
    "condition_notes": ["condition", "condition_notes", "notes", "comments", "state", "quality"],
    "distance_km": ["distance", "distance_km", "km", "dist"],
}

_FEASIBILITY_MAP = {
    # Acquisition
    "stamp_duty_rate_pct": ["stamp_duty", "stamp_duty_rate", "stamp_duty_pct", "stamp_duty_%"],
    "legal_conveyancing": ["legal", "conveyancing", "legal_conveyancing", "solicitor", "legal_fees"],
    "building_pest_inspection": ["building_pest", "bpi", "building_inspection", "pest", "inspection"],
    # Holding
    "finance_interest_rate_annual_pct": ["interest_rate", "finance_rate", "interest", "loan_rate"],
    "finance_lvr_pct": ["lvr", "loan_to_value", "ltv"],
    "council_rates": ["council_rates", "council", "rates"],
    "water_rates": ["water", "water_rates"],
    "insurance": ["insurance", "building_insurance"],
    "land_tax_monthly": ["land_tax", "land_tax_monthly"],
    "utilities": ["utilities", "power", "electricity"],
    # Selling
    "agent_commission_pct": ["commission", "agent_commission", "agent_fee", "agent"],
    "marketing": ["marketing", "advertising", "marketing_cost"],
    "legal_selling": ["legal_selling", "selling_legal", "conveyancing_selling"],
    "styling": ["styling", "staging", "home_staging"],
    # Reno
    "contingency_pct": ["contingency", "contingency_pct", "buffer"],
    # Deal params
    "default_hold_period_months": ["hold_period", "holding_period", "hold", "months"],
    "target_profit_min": ["target_profit", "profit_target", "min_profit"],
    "target_roi_min_pct": ["target_roi", "roi_target", "min_roi"],
    "target_margin_min_pct": ["target_margin", "margin_target", "min_margin"],
}


def _best_match(col_name: str, mapping: dict[str, list[str]]) -> str | None:
    """Find the best target field for a given column name."""
    col = col_name.lower().strip()
    # Pass 1: exact match
    for target, aliases in mapping.items():
        if col in aliases or col == target:
            return target
    # Pass 2: substring match (only for aliases ≥ 4 chars to avoid false positives)
    for target, aliases in mapping.items():
        for alias in aliases:
            if len(alias) >= 4 and (alias in col or col in alias):
                return target
    return None


def _map_row(row: dict[str, Any], mapping: dict[str, list[str]]) -> dict[str, Any]:
    """Map a row's columns to target field names using fuzzy matching."""
    result = {}
    for col, val in row.items():
        target = _best_match(col, mapping)
        if target:
            result[target] = val
        # Keep unmapped columns too — might be useful
    return result


# ═══════════════════════════════════════════════════════════════════════
# CONVERTERS: spreadsheet → specific JSON template
# ═══════════════════════════════════════════════════════════════════════

def convert_dd_checklist(file_bytes: bytes, filename: str) -> dict[str, Any]:
    """
    Convert a spreadsheet to DD checklist JSON.
    Expects columns like: #, Category, Item/Name, Source, Risk, Notes
    """
    rows = read_spreadsheet(file_bytes=file_bytes, filename=filename)
    if not rows:
        raise ValueError("No data found in spreadsheet")

    checklist = []
    for i, row in enumerate(rows, 1):
        mapped = _map_row(row, _DD_CHECKLIST_MAP)
        item = {
            "item_number": mapped.get("item_number", i),
            "category": str(mapped.get("category", "General")),
            "name": str(mapped.get("name", "")),
            "source": str(mapped.get("source", "")),
            "risk_if_fail": _normalize_risk(mapped.get("risk_if_fail", "medium")),
            "notes": str(mapped.get("notes", "")),
        }
        # Skip rows with no name
        if item["name"]:
            checklist.append(item)

    # Re-number
    for i, item in enumerate(checklist, 1):
        item["item_number"] = i

    return {
        "_note": f"Converted from {filename} — {len(checklist)} items",
        "checklist": checklist,
    }


def convert_stores_list(file_bytes: bytes, filename: str) -> dict[str, Any]:
    """
    Convert a spreadsheet to stores list JSON.
    Expects columns like: Category, Store/Name, Notes, Website
    """
    rows = read_spreadsheet(file_bytes=file_bytes, filename=filename)
    if not rows:
        raise ValueError("No data found in spreadsheet")

    stores = []
    for row in rows:
        mapped = _map_row(row, _STORES_MAP)
        store = {
            "category": str(mapped.get("category", "General")),
            "name": str(mapped.get("name", "")),
            "notes": str(mapped.get("notes", "")),
            "website": str(mapped.get("website", "")),
        }
        if store["name"]:
            stores.append(store)

    return {
        "_note": f"Converted from {filename} — {len(stores)} stores",
        "stores": stores,
    }


def convert_comps(file_bytes: bytes, filename: str) -> list[dict[str, Any]]:
    """
    Convert a spreadsheet to comps JSON array.
    Expects columns like: Address, Sold Price, Date, Beds, Baths, Cars, Land, etc.
    """
    rows = read_spreadsheet(file_bytes=file_bytes, filename=filename)
    if not rows:
        raise ValueError("No data found in spreadsheet")

    comps = []
    for row in rows:
        mapped = _map_row(row, _COMPS_MAP)
        comp = {
            "address": str(mapped.get("address", "")),
            "sold_price": _to_int(mapped.get("sold_price", 0)),
            "sold_date": str(mapped.get("sold_date", "")),
            "beds": _to_int(mapped.get("beds", 0)),
            "baths": _to_int(mapped.get("baths", 0)),
            "cars": _to_int(mapped.get("cars", 0)),
            "land_sqm": _to_float(mapped.get("land_sqm", 0)),
            "building_sqm": _to_float(mapped.get("building_sqm", 0)),
            "property_type": str(mapped.get("property_type", "house")),
            "condition_notes": str(mapped.get("condition_notes", "")),
            "distance_km": _to_float(mapped.get("distance_km", 0)),
        }
        if comp["address"]:
            comps.append(comp)

    return comps


def convert_feasibility_template(file_bytes: bytes, filename: str) -> dict[str, Any]:
    """
    Convert a spreadsheet to feasibility template JSON.

    This is more complex because the feasibility template is nested.
    It handles two common layouts:
      1. Two-column key-value layout (Parameter | Value)
      2. Multi-column with categories as rows
    """
    rows = read_spreadsheet(file_bytes=file_bytes, filename=filename)
    if not rows:
        raise ValueError("No data found in spreadsheet")

    # Detect layout: if rows have just 2 meaningful columns, it's key-value
    col_names = list(rows[0].keys()) if rows else []

    # Build a flat dict of all mapped values
    flat = {}
    for row in rows:
        mapped = _map_row(row, _FEASIBILITY_MAP)
        flat.update({k: v for k, v in mapped.items() if v != "" and v is not None})

    # Also try key-value layout: look for columns named "parameter"/"field" + "value"/"amount"
    for row in rows:
        vals = list(row.values())
        keys = list(row.keys())
        # If first column looks like a label and second like a value
        if len(vals) >= 2 and isinstance(vals[0], str):
            label = _clean_col_name(str(vals[0]))
            target = _best_match(label, _FEASIBILITY_MAP)
            if target and vals[1] is not None and vals[1] != "":
                flat[target] = vals[1]

    # Build the nested structure from flat values
    template = {
        "_note": f"Converted from {filename}",
        "acquisition_costs": {
            "stamp_duty_rate_pct": _to_float(flat.get("stamp_duty_rate_pct", 4.5)),
            "stamp_duty_notes": "Converted from spreadsheet",
            "legal_conveyancing": _to_int(flat.get("legal_conveyancing", 2500)),
            "building_pest_inspection": _to_int(flat.get("building_pest_inspection", 800)),
            "other_acquisition": 0,
        },
        "holding_costs_monthly": {
            "finance_interest_rate_annual_pct": _to_float(flat.get("finance_interest_rate_annual_pct", 6.5)),
            "finance_lvr_pct": _to_float(flat.get("finance_lvr_pct", 80)),
            "council_rates": _to_int(flat.get("council_rates", 350)),
            "water_rates": _to_int(flat.get("water_rates", 150)),
            "insurance": _to_int(flat.get("insurance", 250)),
            "land_tax_monthly": _to_int(flat.get("land_tax_monthly", 0)),
            "utilities": _to_int(flat.get("utilities", 100)),
            "other_holding": 0,
        },
        "renovation": {
            "contingency_pct": _to_float(flat.get("contingency_pct", 15)),
            "notes": "Reno budget is entered per deal. Contingency is applied on top.",
        },
        "selling_costs": {
            "agent_commission_pct": _to_float(flat.get("agent_commission_pct", 2.0)),
            "marketing": _to_int(flat.get("marketing", 5000)),
            "legal_selling": _to_int(flat.get("legal_selling", 1500)),
            "styling": _to_int(flat.get("styling", 3000)),
            "other_selling": 0,
        },
        "deal_parameters": {
            "default_hold_period_months": _to_int(flat.get("default_hold_period_months", 6)),
            "target_profit_min": _to_int(flat.get("target_profit_min", 50000)),
            "target_roi_min_pct": _to_float(flat.get("target_roi_min_pct", 15)),
            "target_margin_min_pct": _to_float(flat.get("target_margin_min_pct", 10)),
        },
    }

    return template


# ═══════════════════════════════════════════════════════════════════════
# AUTO-DETECT: figure out which template a spreadsheet is
# ═══════════════════════════════════════════════════════════════════════

def detect_template_type(file_bytes: bytes, filename: str) -> str:
    """
    Auto-detect what kind of template a spreadsheet contains.
    Returns: 'dd_checklist' | 'stores' | 'comps' | 'feasibility' | 'unknown'
    """
    rows = read_spreadsheet(file_bytes=file_bytes, filename=filename)
    if not rows:
        return "unknown"

    cols = set()
    for row in rows:
        cols.update(row.keys())

    col_str = " ".join(cols).lower()

    # DD checklist: has category + name/item + risk/source
    if any(w in col_str for w in ["checklist", "risk", "risk_if_fail"]):
        return "dd_checklist"
    if "source" in col_str and ("category" in col_str or "item" in col_str):
        return "dd_checklist"

    # Stores: has store/supplier + category + website
    if any(w in col_str for w in ["store", "supplier", "vendor"]):
        return "stores"

    # Comps: has address + sold_price/price + beds/baths
    if "address" in col_str and any(w in col_str for w in ["sold", "price", "sale"]):
        return "comps"

    # Feasibility: has commission/stamp_duty/interest type values
    if any(w in col_str for w in ["stamp_duty", "commission", "contingency", "interest_rate"]):
        return "feasibility"

    # Check values for key-value layout (feasibility)
    for row in rows:
        vals = list(row.values())
        if len(vals) >= 2 and isinstance(vals[0], str):
            label = vals[0].lower()
            if any(w in label for w in ["stamp duty", "commission", "interest", "contingency"]):
                return "feasibility"

    return "unknown"


# ── Utility helpers ──────────────────────────────────────────────────

def _normalize_risk(val: Any) -> str:
    val = str(val).lower().strip()
    if val in ("critical", "crit", "1", "very high"):
        return "critical"
    if val in ("high", "2", "h"):
        return "high"
    if val in ("medium", "med", "moderate", "3", "m"):
        return "medium"
    if val in ("low", "4", "l", "minor"):
        return "low"
    return "medium"


def _to_int(val: Any) -> int:
    if isinstance(val, (int, float)):
        return int(val)
    if isinstance(val, str):
        cleaned = val.replace(",", "").replace("$", "").replace("%", "").strip()
        try:
            return int(float(cleaned))
        except (ValueError, TypeError):
            return 0
    return 0


def _to_float(val: Any) -> float:
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        cleaned = val.replace(",", "").replace("$", "").replace("%", "").strip()
        try:
            return float(cleaned)
        except (ValueError, TypeError):
            return 0.0
    return 0.0
